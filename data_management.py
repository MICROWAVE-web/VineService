import csv

import pandas as pd
import xlrd
from openpyxl import Workbook


# pyexcel-xlsx

def load_data(path=r'all_data/data.xls'):
    LoadedData = []
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_by_index(0)
    data = worksheet.cell(5, 5).value
    for x in range(1000):
        try:
            date_check = worksheet.cell(x, 8).value
            if not date_check:
                continue
        except IndexError:
            break
        LoadedData.append({
            'loyal_card_number': worksheet.cell(x, 3).value,
            'phone_number': worksheet.cell(x, 6).value,
            'name': worksheet.cell(x, 5).value,
            'date': worksheet.cell(x, 8).value,
            'articulate': worksheet.cell(x, 9).value,
            'nomn': worksheet.cell(x, 10).value,
            'price': worksheet.cell(x, 13).value,
            'price_with_discount': worksheet.cell(x, 15).value,
        })
        '''for y in range(20):
            data = worksheet.cell(x, y).value
            if data:
                print(f'{x}, {y} - {data}', end='\n\n')'''
    return LoadedData


def mean(numbers):
    return round(sum(numbers)) / max(len(numbers), 1)


def save_ratings(articulate, title, rate, path=r'all_data/ratings.csv', path_xlsx=r'all_data/ratings.xlsx'):
    import csv
    import os
    if not os.path.exists(path):
        with open(path, mode="w+", encoding='utf-8') as w_file:
            file_writer = csv.writer(w_file, delimiter=";", lineterminator="\r")
            file_writer.writerow(["Артикул", "Название", "Оценки", 'Средний балл'])
            # file_writer.writerow([articulate, "0", 0.0])
        #print('Файл создан')
    else:
        # print('Файл существует')
        pass

    df = pd.read_csv(path, sep=';')
    exist = False
    last_index = 0
    for x in range(1000):
        try:
            csv_art = df.iloc[x, 0]
            csv_marks = df.iloc[x, 2]
            csv_avg = df.iloc[x, 3]
        except IndexError:
            last_index = x
            break
        if csv_art == articulate:
            exist = True
            csv_marks += f'{rate}/'
            csv_avg = round(mean(list(map(int, csv_marks.split('/')[:-1]))), 1)
            df.iloc[x, 0] = csv_art
            df.iloc[x, 1] = title
            df.iloc[x, 2] = csv_marks
            df.iloc[x, 3] = csv_avg
    if not exist:
        df.loc[last_index] = [articulate, title, f"{rate}/", float(rate)]
    df.to_csv(path, sep=';', index=False)
    df.to_excel(path_xlsx, index=None, header=True)


def find_client(phone_number=None, card=None):
    if not phone_number:
        phone_number = 'little_bug_fix_:)'
    data = load_data()
    output = []
    for transaction in data:
        clear_phone = str(''.join([str(s) for s in transaction['phone_number'] if s.isdigit()]))
        if phone_number in clear_phone or transaction['loyal_card_number'].split()[0] == card:
            output.append(transaction)
    return output

# save_ratings(articulate='M4047', rate='2')
